#Importaciones
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl as opx
import os


from acciones_sel.login import login
from acciones_sel.entrar_dec import picar_dec, ingresa_per_para_continuar
from acciones_sel.declarar_cero import declarar_cero
from acciones_sel.reporte import esta_en_descargas, mover_descarga, salvar_en_bitacora

# Declaraciones globales
archivo_data = "Data.xlsx"

# ===========================================================================================
# Mis prodecimientos

# Procedimiento que carga los datos
def cargar_datos():
    col_fisica = []
    col_morall = []
    wb = opx.load_workbook(archivo_data)
    wsf = wb["Fisica"]
    wsm = wb["Moral"]
    for i in range(2, wsf.max_row):
        col_fisica.append({
            "rfc"       :   wsf["A" + str(i)].value if wsf["A" + str(i)].value is not None else "",
            "empresa"   :   wsf["B" + str(i)].value if wsf["B" + str(i)].value is not None else "",
            "subido"    :   wsf["C" + str(i)].value if wsf["C" + str(i)].value is not None else "",
            "error"     :   wsf["D" + str(i)].value if wsf["D" + str(i)].value is not None else ""
        })
    for i in range(2, wsm.max_row):
        col_morall.append({
            "rfc"       :   wsm["A" + str(i)].value if wsm["A" + str(i)].value is not None else "",
            "empresa"   :   wsm["B" + str(i)].value if wsm["B" + str(i)].value is not None else "",
            "subido"    :   wsm["C" + str(i)].value if wsm["C" + str(i)].value is not None else "",
            "error"     :   wsm["D" + str(i)].value if wsm["D" + str(i)].value is not None else ""
        })
    wb.close()
    return col_fisica, col_morall

# Procedimientos para mostrar progreso de fisicas
def mostrar_progreso_f(col_fisica):
    if len(col_fisica) == 0:
        print("=> No hay datos para mostrar, cargue datos...")
        return
    print(" +---------------+-----------------+-------+--------------------------------+")
    print(" | RFC           | EMPRESA         | SUBID | ERROR                          |")
    print(" +---------------+-----------------+-------+--------------------------------+")
    for i in range(len(col_fisica)):
        subido = col_fisica[i]['subido'].ljust(5) if col_fisica[i]['subido'] != "" else "No".ljust(5)
        error = col_fisica[i]['error'][:30].ljust(30) if col_fisica[i]['error'] != "" else "Aún no se ha realizado el proceso"[:30].ljust(30)
        print(f" | {col_fisica[i]['rfc']} | {col_fisica[i]['empresa'][:15].ljust(15)} | {subido} | {error} |")
        print(" +---------------+-----------------+-------+--------------------------------+")

# Procedimiento para mostrar progreso de morales  
def mostrar_progreso_m(col_morall):
    if len(col_morall) == 0:
        print("=> No hay datos para mostrar, cargue datos...")
        return
    print(" +---------------+-----------------+-------+--------------------------------+")
    print(" | RFC           | EMPRESA         | SUBID | ERROR                          |")
    print(" +---------------+-----------------+-------+--------------------------------+")
    for i in range(len(col_morall)):
        subido = col_morall[i]['subido'].ljust(5) if col_morall[i]['subido'] != "" else "No".ljust(5)
        error = col_morall[i]['error'][:30].ljust(30) if col_morall[i]['error'] != "" else "Aún no se ha realizado el proceso"[:30].ljust(30)
        print(f" | {col_morall[i]['rfc']}  | {col_morall[i]['empresa'][:15].ljust(15)} | {subido} | {error} |")
        print(" +---------------+-----------------+-------+--------------------------------+")


# Procedimiento para declarar las fisicas
def proceso_fisicas(col_fisicas):
    for i in range(len(col_fisicas)):
        if not (
            (col_fisicas[i]['subido'] == '0' or col_fisicas[i]['subido'] == '') and
            (col_fisicas[i]['error'] == '')
        ): continue

        # Para que vuelva a checar en descargas
        acuse_file = esta_en_descargas(col_fisicas[i]["rfc"])
        if acuse_file:
            mover_descarga(acuse_file)
            col_fisicas[i]['subido'] = '1'
            col_fisicas[i]['error'] = 'Ninguno'
            salvar_en_bitacora(col_fisicas[i], i, 'Fisica')
            continue

        d = webdriver.Chrome()
        d = login(col_fisicas[i]["rfc"], d)

        # Valida si hay error con contraseña
        if type(d).__name__ == 'str':
            col_fisicas[i]['subido'] = '0'
            col_fisicas[i]['error'] = d
            salvar_en_bitacora(col_fisicas[i], i, 'Fisica')
            continue

        # Picar el circulito de al dec si esta gris
        se_pico = picar_dec(col_fisicas[i]["rfc"], d)

        if se_pico :
            d = ingresa_per_para_continuar(d)   # Ingresar periodo y el ejercicio
            d = declarar_cero(d)                # Seleccionar que es dec en ceros
            
            # Si hubo problemas en la descarga
            if type(d).__name__ == 'str':
                col_fisicas[i]['subido'] = '0'
                col_fisicas[i]['error'] = d
                salvar_en_bitacora(col_fisicas[i], i, 'Fisica')
                continue

            acuse_file = esta_en_descargas(col_fisicas[i]["rfc"])
            if acuse_file:
                mover_descarga(acuse_file)
                col_fisicas[i]['subido'] = '1'
                col_fisicas[i]['error'] = 'Ninguno'
                salvar_en_bitacora(col_fisicas[i], i, 'Fisica')
          

# Procedimiento para declarar las morales
def proceso_morales(col_morales):
   for i in range(len(col_morales)):
        if not (
            (col_morales[i]['subido'] == '0' or col_morales[i]['subido'] == '') and
            (col_morales[i]['error'] == '')
        ): continue

        # Para que vuelva a checar en descargas
        acuse_file = esta_en_descargas(col_morales[i]["rfc"])
        if acuse_file:
            mover_descarga(acuse_file)
            col_morales[i]['subido'] = '1'
            col_morales[i]['error'] = 'Ninguno'
            salvar_en_bitacora(col_morales[i], i, 'Fisica')
            continue

        d = webdriver.Chrome()
        d = login(col_morales[i]["rfc"], d)

        # Valida si hay error con contraseña
        if type(d).__name__ == 'str':
            col_morales[i]['subido'] = '0'
            col_morales[i]['error'] = d
            salvar_en_bitacora(col_morales[i], i, 'Fisica')
            continue

        # Picar el circulito de al dec si esta gris
        se_pico = picar_dec(col_morales[i]["rfc"], d)

        if se_pico :
            d = ingresa_per_para_continuar(d)   # Ingresar periodo y el ejercicio
            d = declarar_cero(d,  col_morales[i]["rfc"], True)                # Seleccionar que es dec en ceros
            acuse_file = esta_en_descargas(col_morales[i]["rfc"])
            if acuse_file:
                mover_descarga(acuse_file)
                col_morales[i]['subido'] = '1'
                col_morales[i]['error'] = 'Ninguno'
                salvar_en_bitacora(col_morales[i], i, 'Moral')

# Procedimiento inicial o startport
def main():

    col_fisica = []
    col_morall = []

    while(True):
        print("                 ")
        print("|===============|")
        print("|  DIOT ONLINE  |")
        print("|       1.0     |")
        print("=================")
        print("=> El objetivo de este programa es declarar las diot en automático")
        print("=============================|")
        print("1.- Cargar datos             |")
        print("2.- Mostar progeso fisica    |")
        print("3.- Mostar progeso moral     |")
        print("4.- Seguir con las fisicas   |")
        print("5.- Seguir con las morales   |")
        print("6.- Salir                    |")
        print("=============================|")

        usr_inp = input("=> Selecciona alguna: ")

        if usr_inp == "1":
            os.system("cls")
            col_fisica, col_morall = cargar_datos()

        elif usr_inp == "2":
            mostrar_progreso_f(col_fisica)
        elif usr_inp == "3":
            mostrar_progreso_m(col_morall)
        elif usr_inp == "4":
           col_fisica = proceso_fisicas(col_fisica)
        elif usr_inp == "5":
            col_morall = proceso_morales(col_morall)
        elif usr_inp == "6":
            break
        else:
            print("=> Opcion incorrecta")

   
    # Final del procedimiento
    input("=> El proceso se ha terminado. Presione enter para salir")






















# Llamada al procedimiento inicial
main()