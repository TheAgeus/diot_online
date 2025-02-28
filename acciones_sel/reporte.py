import os
from datetime import datetime
import openpyxl as opx

acuses_path = "C:\\Users\\DIOT\\Desktop\\Diot-online\\acuses\\"
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
archivo_data = "Data.xlsx"

# Procedimiento que verifica si está descargado el acuse
def esta_en_descargas(rfc):
    downloads = os.listdir(downloads_path)

    for d in downloads:
        if 'Acuse.' + rfc in d:
            return d
        
    return False


# Procedimiento que mueve el acuse a su carpeta
def mover_descarga(acuse_file):
    acuses_path = crear_carpetas() 
    dest_path = os.path.join(acuses_path, acuse_file)
    src_path = os.path.join(downloads_path, acuse_file)
    os.rename(src_path, dest_path)
    
# Procedimiento que nos ayuda a crear las carpetas si no existen para lso acuses
def crear_carpetas():
    ahora = datetime.now()
    mes = ahora.month   
    año = ahora.year    
    if mes == 1:
        mes = 12
        año = año - 1
    else:
        mes = mes - 1

    # Ruta de la carpeta que deseas verificar o crear
    ruta_carpeta = os.path.join(acuses_path, str(año), str(mes))

    # Verificar si la carpeta existe
    if not os.path.exists(ruta_carpeta):
        # Si no existe, crear la carpeta
        os.makedirs(ruta_carpeta)
        print(f'Carpeta creada en: {ruta_carpeta}')
    else:
        print(f'La carpeta ya existe en: {ruta_carpeta}')
    return ruta_carpeta

def salvar_en_bitacora(objeto, index, sheet):
    row = index + 2
    wb = opx.load_workbook(archivo_data)
    ws = wb[sheet]
    ws.cell(row,3).value = objeto['subido']
    ws.cell(row,4).value = objeto['error']
    wb.save(archivo_data)
    wb.close()